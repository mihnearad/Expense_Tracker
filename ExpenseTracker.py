import tkinter as tk
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime

class ExpenseTracker:
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.geometry("880x540")
        self.root.resizable(width=False, height=False)
        self.root.title("Expense Tracker")
        self.style = ('Monospace', 18)
        self.labels()
        self.buttons()
        self.root.mainloop()

    # def validate_date(self, input_str):
    #     date_pattern = re.compile(r'^\d{4}\.\d{2}\.\d{2}$')
    #     return bool(date_pattern.match(input_str))

    def validate_numeric(self, input):
        return input.isdigit()

    def labels(self):
        label_frame = tk.Frame(self.root)
        label_frame.grid(row=0, column=0, padx=10, pady=10)

        # Create "Date" label and entry in the label frame
        date_label = tk.Label(label_frame, text="Date:")
        date_label.grid(row=0, column=0, padx=10, pady=10)

        self.date_entry = tk.Entry(label_frame, font=12)
        self.date_entry.grid(row=0, column=1, padx=10, pady=10)

        # validate_date = self.root.register(self.validate_date)
        # self.date_entry.config(validate="key", validatecommand=(validate_date, '%P'))


        # Create "Amount" label and entry in the label frame
        amount_label = tk.Label(label_frame, text="Amount:")
        amount_label.grid(row=1, column=0, padx=10, pady=10)

        self.amount_entry = tk.Entry(label_frame, font=12)
        self.amount_entry.grid(row=1, column=1, padx=10, pady=10)

        validate_numeric = self.root.register(self.validate_numeric)
        self.amount_entry.config(validate="key", validatecommand=(validate_numeric, '%P'))

        # Create "Description" label and entry in the label frame
        description_label = tk.Label(label_frame, text="Description:")
        description_label.grid(row=2, column=0, padx=10, pady=10)

        self.description_entry = tk.Entry(label_frame, font=12)
        self.description_entry.grid(row=2, column=1, padx=10, pady=10)

    def add_expense(self):
        date_value = self.date_entry.get()
        amount_value = self.amount_entry.get()
        description_value = self.description_entry.get()

        # if not date_value or not amount_value:
        #     print("Date and amount are required.")
        #     return

        # Create or load the workbook
        try:
            wb = openpyxl.load_workbook("Expenses.xlsx")
        except FileNotFoundError:
            wb = Workbook()

        # Select the active sheet (create one if it doesn't exist)
        sheet_name = "Expenses"
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        sheet = wb[sheet_name]

        # Add headers if the sheet is empty
        if sheet.max_row == 1:
            sheet.append(["Date", "Amount", "Description"])

        # Format date as YYYY.MM.DD (assuming it's in this format)
        # try:
        #     date_value = datetime.strptime(date_value, "%Y.%m.%d").strftime("%Y.%m.%d")
        # except ValueError:
        #     print("Invalid date format. Please use YYYY.MM.DD.")
        #     return

        # Add a new row with expense details
        sheet.append([date_value, amount_value, description_value])

        # Save the workbook
        wb.save("Expenses.xlsx")
        wb.close()
        print("Expense added successfully.")

        self.date_entry.delete(0, tk.END)
        self.amount_entry.delete(0, tk.END)
        self.description_entry.delete(0, tk.END)


    def buttons(self):
        self.calBudget = tk.Button(self.root, text='Add Expense', width=20, bg='light grey', command=self.add_expense)
        self.calBudget.grid(row=1, column=0, padx=10, pady=10)

if __name__ == '__main__':
    ExpenseTracker()
